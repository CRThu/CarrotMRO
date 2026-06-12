"""
将 txt/csv/xlsx 中的项目名称，匹配数据源中 top 3 最相似的项目，输出为 xlsx。
第一列: 原始名称
第二列: 空列（用户自行填写）
第三~五列: 三个最高匹配项目
"""

import sys
import json
import csv
from pathlib import Path
import openpyxl
from rapidfuzz import process, fuzz

DEFAULT_DATA = "data/aa.json"
DEFAULT_OUTPUT = "匹配结果.xlsx"
TOP_N = 3


def load_items(data_path: str) -> list[str]:
    with open(data_path, encoding="utf-8") as f:
        return json.load(f)["items"]


def read_names(input_path: str) -> list[str]:
    path = Path(input_path)
    if not path.exists():
        print(f"错误: 文件不存在 - {input_path}")
        sys.exit(1)

    suffix = path.suffix.lower()
    names: list[str] = []

    if suffix == ".csv":
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if row and row[0].strip():
                    names.append(row[0].strip())
    elif suffix == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # 跳过表头
            if row and row[0] and str(row[0]).strip():
                names.append(str(row[0]).strip())
        wb.close()
    else:
        # 默认当作 txt，每行一个名称
        with open(path, encoding="utf-8") as f:
            for line in f:
                stripped = line.strip()
                if stripped:
                    names.append(stripped)

    return names


def match_names(names: list[str], items: list[str], n: int = TOP_N) -> list[dict]:
    results = []
    for name in names:
        matches = process.extract(name, items, scorer=fuzz.token_set_ratio, limit=n)
        top3 = [m[0] for m in matches]
        # 不够 3 个用空字符串补位
        while len(top3) < n:
            top3.append("")
        results.append({
            "original": name,
            "matches": top3,
        })
    return results


def write_xlsx(results: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "匹配结果"

    # 表头
    headers = ["项目名称", "用户选择", "匹配1", "匹配2", "匹配3"]
    ws.append(headers)

    for r in results:
        row = [r["original"], ""] + r["matches"]
        ws.append(row)

    # 自动调整列宽（简单适配）
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                # 中文字符算 2 个宽度
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 60)

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("用法: uv run python match.py <输入文件.txt|csv|xlsx> [-d <数据文件>] [-o <输出.xlsx>]")
        print("示例: uv run python match.py names.txt")
        print("      uv run python match.py names.csv -d data/standard.json")
        print("      uv run python match.py 已有文件.xlsx -o 匹配结果.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    data_path = DEFAULT_DATA
    output_path = DEFAULT_OUTPUT

    if "-d" in sys.argv:
        idx = sys.argv.index("-d")
        if idx + 1 < len(sys.argv):
            data_path = sys.argv[idx + 1]

    if "-o" in sys.argv:
        idx = sys.argv.index("-o")
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    # 1. 读取数据源
    print(f"加载数据源: {data_path}")
    items = load_items(data_path)
    print(f"  共 {len(items)} 条项目")

    # 2. 读取输入名称
    names = read_names(input_path)
    print(f"读取输入: {input_path}")
    print(f"  共 {len(names)} 个名称")

    # 3. 匹配
    print("匹配中...")
    results = match_names(names, items)

    # 4. 输出 xlsx
    write_xlsx(results, output_path)
    print(f"完成! 结果已保存到: {output_path}")


if __name__ == "__main__":
    main()
