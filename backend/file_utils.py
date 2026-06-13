import json
import csv
from pathlib import Path
import openpyxl
import sys

def load_items(data_path: str) -> list[str]:
    with open(data_path, encoding="utf-8") as f:
        return json.load(f)["items"]

def read_names(input_path: str) -> list[str]:
    path = Path(input_path)
    if not path.exists():
        print(f"错误: 文件不存在 - {input_path}")
        sys.exit(1)

    suffix = path.suffix.lower()
    names: list[str] = []

    if suffix == ".csv":
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if row and row[0].strip():
                    names.append(row[0].strip())
    elif suffix == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # 跳过表头
            if row and row[0] and str(row[0]).strip():
                names.append(str(row[0]).strip())
        wb.close()
    else:
        # 默认当作 txt，每行一个名称
        with open(path, encoding="utf-8") as f:
            for line in f:
                stripped = line.strip()
                if stripped:
                    names.append(stripped)

    return names

def write_xlsx(results: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "匹配结果"

    # 表头
    headers = ["项目名称", "用户选择", "匹配1", "匹配2", "匹配3"]
    ws.append(headers)

    for r in results:
        row = [r["original"], ""] + r["matches"]
        ws.append(row)

    # 自动调整列宽（简单适配）
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                # 中文字符算 2 个宽度
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 60)

    wb.save(output_path)
