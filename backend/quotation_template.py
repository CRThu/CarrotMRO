"""报价单 Excel 模板解析器。

占位符采用命名空间前缀区分循环层级:
  group.xxx  - 组级占位符（每组输出一次）
    {group.name}     - 组名
    {group.num}      - 组内序号（从 1 开始）

  item.xxx   - 数据级占位符（每个 item 输出一次）
    {item.name}      - 项目名称
    {item.unit}      - 单位
    {item.quantity}  - 数量
    {item.unit_price} - 单价

  row 相关  - 行号引用（导出时替换为实际行号）
    {row}            - 当前行号
    {row-N}          - 当前行号减 N
    {row+N}          - 当前行号加 N

模板结构:
  标题行（无占位符，保持不动）
  {group.name} 行（每组复制一次，合并单元格）
  {item.xxx} 行（每个 item 复制一次）
  页脚行（无占位符，保持不动，含 {row} 公式会更新行号）
"""

import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from copy import copy

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter


PLACEHOLDER_RE = re.compile(r"\{([^}]+)\}")


@dataclass
class Placeholder:
    """占位符信息。"""
    name: str          # 完整名称，如 'group.name', 'item.unit', 'row-2'
    row: int           # 所在行号 (1-based)
    col: int           # 所在列号 (1-based)
    raw: str           # 原始文本，如 '{group.name}'


@dataclass
class TemplateRow:
    """模板行信息。"""
    row: int                          # 行号 (1-based)
    kind: str                         # 'header' | 'group' | 'data'
    placeholders: list[Placeholder] = field(default_factory=list)
    has_formula: bool = False         # 是否包含 Excel 公式


@dataclass
class QuotationTemplate:
    """解析后的报价单模板。

    导出时只需要 group_row 和 data_row 两个模板行:
    - group_row: 组名行模板，包含 group.* 占位符，每组复制一次
    - data_row: 数据行模板，包含 item.* 占位符，每个 item 复制一次

    其他行（header、empty）保持不动，不参与数据填充。
    包含 {row} 占位符的公式在导出时统一更新行号引用。
    """
    group_row: int | None             # 组名模板行
    data_row: int | None              # 数据模板行
    group_merge_range: str | None     # 组名行的合并单元格范围（如 'A5:F5'）
    placeholders: dict[str, Placeholder] = field(default_factory=dict)


def load_template(content: bytes) -> QuotationTemplate:
    """从 Excel 文件内容加载并解析模板。"""
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    rows: list[TemplateRow] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        template_row = _parse_row(row_idx, row)
        rows.append(template_row)

    if not rows:
        raise ValueError("模板为空")

    template = _build_template(rows)

    # 检测组名行的合并单元格范围
    if template.group_row:
        for merge in ws.merged_cells.ranges:
            if merge.min_row == template.group_row:
                template.group_merge_range = str(merge)
                break

    return template


def load_template_from_file(path: str | Path) -> QuotationTemplate:
    """从文件路径加载模板。"""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"模板文件不存在: {p}")
    return load_template(p.read_bytes())


def _parse_row(row_idx: int, cells: tuple[Cell, ...]) -> TemplateRow:
    """解析单行，识别占位符和公式。"""
    placeholders = []
    has_formula = False
    non_empty_count = 0

    for col_idx, cell in enumerate(cells, start=1):
        val = cell.value
        if val is None:
            continue

        val_str = str(val).strip()
        if not val_str:
            continue

        non_empty_count += 1

        if val_str.startswith("="):
            has_formula = True

        for m in PLACEHOLDER_RE.finditer(val_str):
            placeholders.append(Placeholder(
                name=m.group(1),
                row=row_idx,
                col=col_idx,
                raw=m.group(0),
            ))

    kind = _classify_row(row_idx, non_empty_count, placeholders)

    return TemplateRow(
        row=row_idx,
        kind=kind,
        placeholders=placeholders,
        has_formula=has_formula,
    )


def _classify_row(
    row_idx: int,
    non_empty_count: int,
    placeholders: list[Placeholder],
) -> str:
    """根据占位符命名空间判断行类型。

    分类规则:
    1. 空行 → 'header'
    2. 有 item.* 占位符 → 'data'（数据行，每个 item 输出一次）
       注意: 数据行可能同时有 group.num，优先级高于纯 group 行
    3. 只有 group.* 占位符（无 item.*）→ 'group'（组名行，每组输出一次）
    4. 其他非空行 → 'header'（标题、页脚等固定内容）
    """
    if non_empty_count == 0:
        return "header"

    namespaces = {p.name.split(".")[0] for p in placeholders if "." in p.name}

    # 有 item.* 就是数据行（即使同时有 group.num）
    if "item" in namespaces:
        return "data"

    # 只有 group.* 没有 item.* 才是组名行
    if "group" in namespaces:
        return "group"

    return "header"


def _build_template(rows: list[TemplateRow]) -> QuotationTemplate:
    """从解析后的行构建模板对象。

    找到 group_row 和 data_row 后，收集所有占位符信息。
    导出时只需要这两个模板行的位置，其他行保持不动。
    """
    group_row = None
    data_row = None
    all_placeholders: dict[str, Placeholder] = {}

    for tr in rows:
        if tr.kind == "group":
            if group_row is not None:
                raise ValueError(f"模板中存在多个组名模板行: 第{group_row}行和第{tr.row}行")
            group_row = tr.row
        elif tr.kind == "data":
            if data_row is not None:
                raise ValueError(f"模板中存在多个数据模板行: 第{data_row}行和第{tr.row}行")
            data_row = tr.row

        for p in tr.placeholders:
            all_placeholders[p.name] = p

    if group_row is None:
        raise ValueError("模板中未找到组名模板行（需包含 {group.name} 占位符）")
    if data_row is None:
        raise ValueError("模板中未找到数据模板行（需包含 {item.name} 占位符）")

    return QuotationTemplate(
        group_row=group_row,
        data_row=data_row,
        group_merge_range=None,
        placeholders=all_placeholders,
    )


def get_template_info(template: QuotationTemplate) -> dict:
    """获取模板的摘要信息（用于 API 返回）。"""
    return {
        "group_row": template.group_row,
        "data_row": template.data_row,
        "group_merge_range": template.group_merge_range,
        "placeholders": {
            name: {"row": p.row, "col": p.col, "raw": p.raw}
            for name, p in template.placeholders.items()
        },
    }


def get_template_columns(template: QuotationTemplate) -> list[str]:
    """提取模板中所有可用的数据列名。

    从 {item.xxx} 占位符中提取 xxx 作为可用列名。

    Returns:
        列名列表，如 ["物料名称", "数量", "单位", "单价"]
    """
    columns = []
    for name in template.placeholders:
        if name.startswith("item."):
            col_name = name[5:]
            if col_name not in columns:
                columns.append(col_name)
    return columns


def export_quotation(
    template_content: bytes,
    groups: list[dict],
) -> bytes:
    """将报价单数据填充到模板，生成 Excel 文件。

    Args:
        template_content: 模板 Excel 文件的二进制内容
        groups: 分组数据列表，每个元素格式:
            {
                "name": "组名",
                "items": [
                    {"name": "项目名称", "unit": "单位", "quantity": "数量", "unit_price": "单价"},
                ]
            }

        items 中的 key 为纯字段名（不带 item. 前缀），导出时自动映射到模板占位符。

    Returns:
        生成的 Excel 文件二进制内容
    """
    template = load_template(template_content)
    wb = openpyxl.load_workbook(BytesIO(template_content))
    ws = wb.active

    # 保存模板行号（插入行前的原始位置，用于样式复制）
    src_group_row = template.group_row
    src_data_row = template.data_row

    # 过滤掉空组
    groups = [g for g in groups if g.get("items")]

    # 保存模板行的原始公式（在插入行之前保存，避免复制已处理的公式）
    data_formulas: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=src_data_row, column=col)
        if cell.value and str(cell.value).startswith("="):
            data_formulas[col] = str(cell.value)

    # 计算需要插入的行数: 每个组1行组名 + N行数据，减去已有的2行模板
    total_items = sum(len(g.get("items", [])) for g in groups)
    total_group_rows = len(groups)
    total_new_rows = total_group_rows + total_items
    rows_to_insert = total_new_rows - 2  # 减去已有的 group_row 和 data_row

    if rows_to_insert > 0:
        ws.insert_rows(src_data_row + 1, rows_to_insert)

    # 填充数据: 遍历每个组，先填组名行，再逐行填数据
    # current_row 从 src_group_row 开始，逐行递增
    current_row = src_group_row
    for group_idx, group in enumerate(groups):
        # 填充组名行: 从模板原始位置复制样式，填充到当前行
        _copy_row_style(ws, src_group_row, current_row)
        _fill_row_placeholder(ws, template, current_row, {
            "group.name": group.get("name", ""),
            "group.num": str(group_idx + 1),
        }, is_group_row=True)
        current_row += 1

        # 填充数据行: 每行从模板原始位置复制样式，替换占位符
        # 输入 key 为纯字段名（如 name），自动映射到模板占位符（如 item.name）
        items = group.get("items", [])
        for item_idx, item in enumerate(items):
            _copy_row_style(ws, src_data_row, current_row)
            item_values = {"group.num": str(item_idx + 1)}
            for k, v in item.items():
                item_values[f"item.{k}"] = v
            _fill_row_placeholder(ws, template, current_row, item_values, is_group_row=False)
            # 替换数据行公式中的行号（如 {row} → 当前行号）
            for col, formula_template in data_formulas.items():
                cell = ws.cell(row=current_row, column=col)
                cell.value = _replace_row_refs(formula_template, current_row)
            current_row += 1

    # 更新页脚公式中的行号占位符
    # 页脚行在 src_data_row + rows_to_insert 之后
    footer_start = src_data_row + rows_to_insert + 1 if rows_to_insert > 0 else src_data_row + 1
    for row in range(footer_start, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and str(cell.value).startswith("=") and "{" in str(cell.value):
                cell.value = _replace_row_refs(str(cell.value), row)

    # 为每个组创建合并单元格（如果有）
    if template.group_merge_range:
        # 先取消模板组名行的原始合并单元格
        for merge in list(ws.merged_cells.ranges):
            if merge.min_row == src_group_row:
                ws.unmerge_cells(str(merge))
                break

        # 解析模板合并范围（如 "A5:F5"）
        merge_str = template.group_merge_range
        parts = merge_str.split(":")
        start_col_str = re.match(r"[A-Za-z]+", parts[0]).group()
        end_col_str = re.match(r"[A-Za-z]+", parts[1]).group()
        start_col = _col_str_to_num(start_col_str.upper())
        end_col = _col_str_to_num(end_col_str.upper())

        # 为每个组创建合并单元格
        current_row = src_group_row
        for group in groups:
            min_cell = f"{get_column_letter(start_col)}{current_row}"
            max_cell = f"{get_column_letter(end_col)}{current_row}"
            ws.merge_cells(f"{min_cell}:{max_cell}")
            current_row += 1 + len(group.get("items", []))

    # 保存到 bytes
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _col_str_to_num(col_str: str) -> int:
    """将 Excel 列字母转换为列号（A=1, B=2, ..., Z=26, AA=27）。"""
    result = 0
    for c in col_str:
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result


def _copy_row_style(ws, template_row: int, target_row: int):
    """复制模板行的样式到目标行。

    包括: 字体、边框、填充色、数字格式、保护、对齐方式。
    不复制单元格值（值由后续占位符填充）。
    """
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=template_row, column=col)
        dst_cell = ws.cell(row=target_row, column=col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)


def _fill_row_placeholder(ws, template: QuotationTemplate, row: int, values: dict, is_group_row: bool):
    """填充行中的占位符。

    Args:
        ws: 工作表
        template: 模板对象
        row: 当前要填充的行号
        values: 占位符名称 → 值的映射（如 {"group.name": "xxx"}）
        is_group_row: 是否是组名行
    """
    template_row = template.group_row if is_group_row else template.data_row

    for name, placeholder in template.placeholders.items():
        # 只处理当前模板行的占位符
        if placeholder.row != template_row:
            continue
        # 跳过行号占位符（{row} 等由 _replace_row_refs 处理）
        if name.startswith("row"):
            continue

        cell = ws.cell(row=row, column=placeholder.col)
        val = values.get(name, "")
        if val:
            cell.value = val


def _replace_row_refs(formula: str, current_row: int) -> str:
    """替换公式中的行号占位符。

    {row}    → current_row
    {row-1}  → current_row - 1
    {row+1}  → current_row + 1
    """
    def replacer(match):
        ref = match.group(1)
        if ref == "row":
            return str(current_row)
        elif ref.startswith("row-"):
            n = int(ref[4:])
            return str(current_row - n)
        elif ref.startswith("row+"):
            n = int(ref[4:])
            return str(current_row + n)
        return match.group(0)

    return PLACEHOLDER_RE.sub(replacer, formula)


def import_quotation(
    template_content: bytes,
    excel_content: bytes,
) -> list[dict]:
    """从 Excel 文件中提取报价单数据。

    根据模板结构识别组名行和数据行，提取数据并按组分组返回。

    Args:
        template_content: 模板 Excel 文件的二进制内容（用于识别行结构）
        excel_content: 要导入的 Excel 文件的二进制内容

    Returns:
        分组数据列表，格式同 export_quotation 的 groups 参数:
        [
            {
                "name": "组名",
                "items": [
                    {"name": "项目名称", "unit": "单位", ...},
                ]
            }
        ]
    """
    template = load_template(template_content)
    template_placeholders = template.placeholders

    # 建立列号 → 占位符名称的映射（只关心 group_row 和 data_row）
    group_col_map: dict[int, str] = {}  # col → "group.name"
    data_col_map: dict[int, str] = {}   # col → "item.xxx"

    for name, ph in template_placeholders.items():
        if name.startswith("row"):
            continue
        if ph.row == template.group_row:
            group_col_map[ph.col] = name
        elif ph.row == template.data_row:
            data_col_map[ph.col] = name

    # 加载要导入的 Excel
    wb = openpyxl.load_workbook(BytesIO(excel_content))
    ws = wb.active

    groups: list[dict] = []
    current_group: dict | None = None
    total_rows = ws.max_row

    def _row_has_data(row_idx: int) -> bool:
        """检查某行的数据列（item.*）是否有值"""
        for col, name in data_col_map.items():
            if not name.startswith("item."):
                continue
            cell = ws.cell(row=row_idx, column=col)
            val = cell.value
            if val is not None:
                val_str = str(val).strip()
                if val_str and not val_str.startswith("="):
                    return True
        return False

    for row_idx in range(1, total_rows + 1):
        is_group_row = False
        if group_col_map and row_idx >= template.group_row:
            group_val = None
            for col, name in group_col_map.items():
                cell = ws.cell(row=row_idx, column=col)
                val = cell.value
                if val is not None:
                    group_val = str(val).strip()

            # 组名行: group.name 有值 + 本行无数据 + 下一行有数据（排除页脚行）
            if group_val and not _row_has_data(row_idx):
                next_has_data = _row_has_data(row_idx + 1) if row_idx < total_rows else False
                if next_has_data:
                    is_group_row = True
                    current_group = {
                        "name": group_val,
                        "items": [],
                    }
                    groups.append(current_group)

        # 尝试识别为数据行: 仅检查 item.* 列
        if not is_group_row and data_col_map and current_group is not None:
            item: dict[str, str] = {}
            has_value = False
            for col, name in data_col_map.items():
                if not name.startswith("item."):
                    continue
                cell = ws.cell(row=row_idx, column=col)
                val = cell.value
                if val is not None:
                    val_str = str(val).strip()
                    if val_str.startswith("="):
                        continue
                    if val_str:
                        field_name = name.split(".", 1)[1]
                        item[field_name] = val_str
                        has_value = True

            # 如果数据行有值，添加到当前组
            if has_value and item:
                current_group["items"].append(item)

    return groups
