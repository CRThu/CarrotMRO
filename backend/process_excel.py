import openpyxl
import json
import argparse
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent

# ═══════════════════════════════════════════
# 命令行参数
# ═══════════════════════════════════════════
parser = argparse.ArgumentParser(description="从标准项目Excel提取完整数据行")
parser.add_argument("-i", "--input", default=str(ROOT_DIR / "project-template" / "standard.xlsx"),
                    help="输入Excel文件路径 (默认: project-template/standard.xlsx)")
parser.add_argument("-o", "--output", default=None,
                    help="输出JSON文件路径 (默认: data/<同名>.json)")
args = parser.parse_args()

input_path = args.input
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# ═══════════════════════════════════════════
# 配置区（关键词匹配 $ 内容）
# ═══════════════════════════════════════════

# 匹配关键词：用于判断数据完整性（非空检查）
# 匹配逻辑：$内容 包含 此关键词的列，都需非空
REQUIRED_KEYWORDS = ["名称", "单位", "单价"]

# 导出列关键词：哪一列作为 "名称" 导出到 JSON
NAME_KEYWORD = "名称"

# ═══════════════════════════════════════════

# 读取第1行（标记行）和第2行（表头行）
marker_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

print("Marker row:", list(marker_row))
print("Header row:", list(header_row))

# 解析所有 $ 列：[(col_idx, $name, header_name), ...]
dollar_columns = []
for col_idx, val in enumerate(marker_row):
    if val is not None and str(val).startswith("$"):
        dollar_name = str(val).lstrip("$").strip()
        header_name = str(header_row[col_idx]).strip() if col_idx < len(header_row) and header_row[col_idx] is not None else dollar_name
        dollar_columns.append((col_idx, dollar_name, header_name))

DATA_COLS = [h for _, _, h in dollar_columns]
print(f"数据列 (DATA_COLS): {DATA_COLS}")
print(f"匹配关键词 (REQUIRED_KEYWORDS): {REQUIRED_KEYWORDS}")
print(f"导出列关键词 (NAME_KEYWORD): {NAME_KEYWORD}")

# 根据关键词匹配 $ 内容，找到匹配列索引
required_col_indices = set()
for col_idx, dollar_name, header_name in dollar_columns:
    for kw in REQUIRED_KEYWORDS:
        if kw in dollar_name:
            required_col_indices.add(col_idx)
            print(f"  匹配: ${dollar_name} (表头: {header_name}) 包含关键词 '{kw}' -> 列{col_idx}")
            break

required_col_indices = sorted(required_col_indices)
print(f"匹配列索引 (用于非空检查): {required_col_indices}")

# 找到导出列（名称列）
name_col_idx = None
name_header = None
for col_idx, dollar_name, header_name in dollar_columns:
    if NAME_KEYWORD in dollar_name:
        name_col_idx = col_idx
        name_header = header_name
        print(f"导出列: ${dollar_name} (表头: {header_name}) -> 列{col_idx}")
        break

if name_col_idx is None:
    print("Error: 未找到匹配导出关键词的列!")
    exit(1)

# 遍历数据行（从第3行开始）
results = []
data_start_row = 3
for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=True), data_start_row):
    values = list(row)

    # 检查所有匹配列是否非空
    all_complete = True
    for idx in required_col_indices:
        if idx >= len(values) or values[idx] is None or str(values[idx]).strip() == "":
            all_complete = False
            break

    if all_complete:
        name_val = values[name_col_idx] if name_col_idx < len(values) else None
        if name_val is not None and str(name_val).strip():
            results.append(str(name_val).strip())
            print(f"Row {row_idx}: OK 完整 -> {name_val}")
        else:
            print(f"Row {row_idx}: OK 完整但名称列为空 -> 跳过")
    else:
        print(f"Row {row_idx}: NG 不完整 -> 跳过")

output = {"items": results}

# 确定输出路径
if args.output:
    output_path = Path(args.output)
else:
    output_path = ROOT_DIR / "data" / f"{Path(input_path).stem}.json"
output_path.parent.mkdir(parents=True, exist_ok=True)

with open(output_path, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=None)

print(f"\n{'='*50}")
print(f"输入文件: {input_path}")
print(f"输出文件: {output_path}")
print(f"数据列 (DATA_COLS): {DATA_COLS}")
print(f"匹配关键词: {REQUIRED_KEYWORDS}")
print(f"匹配列索引: {required_col_indices}")
print(f"导出列: {name_header} (列{name_col_idx})")
print(f"有效条目: {len(results)}")
print(f"{'='*50}")
